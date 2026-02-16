# TUGAS STATISTIKA ANALISIS DATA

import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

x = {
    "nama": "Chris",
    "kelas": "2025 TIG",
    "Nim": 25051204061
}

y = {
    "No": [1,2,3,4,5,6,7,8,9,10],
    "Nama Logam": [
        "Besi",
        "Emas",
        "Perak",
        "Tembaga",
        "Platina",
        "Aluminium",
        "Timah",
        "Nikel",
        "Titanium",
        "Paladium"
    ],
    "Status Kelangkaan": [
        "Tidak Langka",
        "Langka",
        "Tidak Langka",
        "Tidak Langka",
        "Langka",
        "Tidak Langka",
        "Tidak Langka",
        "Tidak Langka",
        "Langka",
        "Langka"
    ],
    "Ketahanan Korosi": [
        "Rendah",
        "Tinggi",
        "Sedang",
        "Sedang",
        "Tinggi",
        "Sedang",
        "Rendah",
        "Tinggi",
        "Tinggi",
        "Tinggi"
    ],
    "Titik Lebur(C)": [
        1538,
        1064,
        962,
        1085,
        1768,
        660,
        232,
        1455,
        1668,
        1555
    ],
    "Harga per Gram": [
        60,
        2938000,
        48000,
        6000,
        1200000,
        25,
        50,
        10000,
        45000,
        500000
    ]
}

tugas = pd.DataFrame(y)

print("\nTUGAS STATISTIKA ANALISIS DATA\n")
print(f"Nama    : {x['nama']}\nkelas  : {x['kelas']}\nNim   : {x['Nim']}\n")

print("---------- DATA BATUAN LOGAM ----------\n")
print(tugas.to_string(index=False, justify="center"))

folder_script = os.path.dirname(os.path.abspath(__file__))
tugas2 = os.path.join(folder_script, "Chris 25051204061.xlsx")

tugas.to_excel(tugas2, index=False)

wb = load_workbook(tugas2)
ws = wb.active

header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
body_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")

font_body = Font(name="Times New Roman",)
font_header = Font(name="Times New Roman", bold=True, color="FFFFFF")
thin_border = Border(
    left=Side(style="thin", color="000000"),
    right=Side(style="thin", color="000000"),
    top=Side(style="thin", color="000000"),
    bottom=Side(style="thin", color="000000")
)

for cell in ws[1]:
    cell.fill = header_fill
    cell.font = font_header
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.border = thin_border

for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
    for cell in row:
        cell.fill = body_fill
        cell.font = font_body
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = thin_border

for row in ws.iter_rows(min_row=2, min_col=6, max_col=6):
    for cell in row:
        cell.number_format = '"Rp"#,##0'

for column in ws.columns:
    max_length = 0
    column_letter = column[0].column_letter
    for cell in column:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    ws.column_dimensions[column_letter].width = max_length + 2

wb.save(tugas2)
penjelasan ="""
ANALISI DATA 

TIPE DATA :

        1.Kuantitatif, data numerik yang dapat dihitung dan diukur 
        Contoh dalam tabel : Titik lebur & Harga per Gram

        2:Kualitatif,data yang menggmbarkan karakteristik yang tidak dapat diukur dengan angka
        Contoh dalam tabel : Nama logam, Status Kelangkaan, Ketahanan Korosi

SKALA PENGUKURAN:

        1.Nominal : Hanya menggambarkan karakteristik tanpa ada urutan khusus
        Contoh dalam tabel : Nama Logam

        2.Ordinal : Data dengan urutan yang jelas namum perbedaannya tidak dapat diukur dengan pasti
        Contoh dalam tabel : Tingkat korosi

        3.Interval, Data dengan interval seragam memungkinkan perbandingan matematis
        Contoh dalam tabel : Titik Lebur

        4.Rasio : Memiliki Interval yang seragam dan memiliki titik 0 absolut
        Contoh dalam tabel : Harga per Gram"""

print(penjelasan)
os.startfile(tugas2)
